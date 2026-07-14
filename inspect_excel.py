import json
from pathlib import Path
import openpyxl

files = [
    'ASK Call - Suivi de ventes MINT (réponses) (1).xlsx',
    'tableau statut brut.xlsx'
]

res = []
for fname in files:
    p = Path(fname)
    if not p.exists():
        res.append({'file': fname, 'error': 'missing'})
        continue
    wb = openpyxl.load_workbook(p, data_only=True)
    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        rows = []
        for row in ws.iter_rows(min_row=1, max_row=7, values_only=True):
            rows.append([cell for cell in row])
        sheets.append({'name': name, 'rows': rows})
    res.append({'file': fname, 'sheets': sheets})
print(json.dumps(res, ensure_ascii=False, indent=2))
